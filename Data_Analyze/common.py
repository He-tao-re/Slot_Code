import os
import json
import openpyxl

import Data_Analyze.Variable as Variable

def file_path(file_dir):
    files_path = []
    for root, dirs, files in os.walk(file_dir):
        # print(root) #当前目录路径
        # print(dirs) #当前路径下所有子目录
        # print(files) #当前路径下所有非目录子文件

        for file in files:
            files_path.append(root + '/' + file)

        # print(files_path)
    return files_path

def sort_data(original_file_name,sort_file_name):
    #文档路径 + 储存路径
    user_game_data = open(original_file_name, 'r', newline='')

    data_save = {}
    time_list = []


    for line in user_game_data:
        spin_data = json.loads(line)
        uid = spin_data["data"]['uid']
        create_Time = spin_data['data']['serverTime']
        time_list.append(create_Time)
        data_save[create_Time] = line

    time_list.sort()

    sort_file = open(sort_file_name, 'a+', newline='')

    for time_point in time_list:
        sort_file.write(data_save[time_point])

    sort_file.close()

def rtp_depart(summary_store,subdivide_data_file):

    try:
        os.remove(subdivide_data_file)
        wb = openpyxl.Workbook()
        wb.save(subdivide_data_file)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        wb.save(subdivide_data_file)

    rtp_list = [Variable.RTP120, Variable.RTP110, Variable.RTP100, Variable.RTP98, Variable.RTP96, Variable.RTP94,
                Variable.RTP90, Variable.RTP85, Variable.RTP80, Variable.RTP70, Variable.NOFEATURE, Variable.REWARD, Variable.All_Spin]

    wb = openpyxl.load_workbook(subdivide_data_file)

    for rtpType in rtp_list:
        sht = wb.create_sheet(rtpType)


        sht.cell(1,1,"uid")
        sht.cell(1,2,"spinTimes")
        sht.cell(1,3,"累计bet")
        sht.cell(1,4,"累计win")
        sht.cell(1,5,"累计倍数")
        sht.cell(1,6,"rtp")

        num = 1
        user_data_file = open(summary_store, 'r', newline='')

        for user_data in user_data_file:
            user_data_dic = json.loads(user_data)

            num += 1


            uid = list(user_data_dic.keys())[0]
            spinTimes = user_data_dic[uid][rtpType]["Spin"]
            accBet = user_data_dic[uid][rtpType]['accumulative_bet']
            accWin = user_data_dic[uid][rtpType]['accumulative_win']
            accMul = user_data_dic[uid][rtpType]['accumulative_mul']
            rtp = user_data_dic[uid][rtpType]['RTP']
            # print(num)

            sht.cell(num, 1, int(uid))
            sht.cell(num, 2, spinTimes)
            sht.cell(num, 3, accBet)
            sht.cell(num, 4, accWin)
            sht.cell(num, 5, accMul)
            sht.cell(num, 6, rtp)

        excel_analysis_data(sht)

        user_data_file.close()
    wb.save(subdivide_data_file)

def data_summary(summary_store,game_id,summary_data):
    rtp_list = [Variable.All_Spin]

    wb = openpyxl.load_workbook(summary_data)
    sht = wb.create_sheet(str(game_id))

    for rtpType in rtp_list:

        sht.cell(1,1,"uid")
        sht.cell(1,2,"spinTimes")
        sht.cell(1,3,"累计bet")
        sht.cell(1,4,"累计win")
        sht.cell(1,5,"累计倍数")
        sht.cell(1,6,"rtp")

        num = 1
        user_data_file = open(summary_store, 'r', newline='')

        for user_data in user_data_file:
            user_data_dic = json.loads(user_data)

            num += 1

            cell = 'a' + str(num)

            uid = list(user_data_dic.keys())[0]
            spinTimes = user_data_dic[uid][rtpType]["Spin"]
            accBet = user_data_dic[uid][rtpType]['accumulative_bet']
            accWin = user_data_dic[uid][rtpType]['accumulative_win']
            accMul = user_data_dic[uid][rtpType]['accumulative_mul']
            rtp = user_data_dic[uid][rtpType]['RTP']
            # print(num)

            sht.cell(num, 1, int(uid))
            sht.cell(num, 2, spinTimes)
            sht.cell(num, 3, accBet)
            sht.cell(num, 4, accWin)
            sht.cell(num, 5, accMul)
            sht.cell(num, 6, rtp)

        excel_analysis_data(sht)

        user_data_file.close()
    wb.save(summary_data)

def excel_analysis_data(sht):
    sht.cell(1, 14, '玩家spin分布数据')
    sht.cell(2, 14, '')
    sht.cell(3, 14, '[0]')
    sht.cell(4, 14, '(0,10]')
    sht.cell(5, 14, '(10,50]')
    sht.cell(6, 14, '(50,100]')
    sht.cell(7, 14, '(100,500]')
    sht.cell(8, 14, '(500,max]')
    sht.cell(9, 14, '')
    sht.cell(10, 14, '人均spin')
    sht.cell(11, 14, '人均spin（Spin >= 100）')
    sht.cell(12, 14, '深度体验率')
    sht.cell(13, 14, '')
    sht.cell(14, 14, '玩家RTP分布数据(区间)')
    sht.cell(15, 14, '(0,0.3]')
    sht.cell(16, 14, '(0.3,0.4]')
    sht.cell(17, 14, '(0.4,0.5]')
    sht.cell(18, 14, '(0.5,0.6]')
    sht.cell(19, 14, '(0.6,0.7]')
    sht.cell(20, 14, '(0.7,0.8]')
    sht.cell(21, 14, '(0.8,0.9]')
    sht.cell(22, 14, '(0.9,1]')
    sht.cell(23, 14, '(1,1.1]')
    sht.cell(24, 14, '(1.1,1.2]')
    sht.cell(25, 14, '(1.2,1.5]')
    sht.cell(26, 14, '(1.5,2]')
    sht.cell(27, 14, '(2,max]')
    sht.cell(28, 14, '')
    sht.cell(29, 14, '人均RTP')
    sht.cell(30, 14, '关卡RTP')
    sht.cell(31, 14, '')
    sht.cell(32, 14, '玩家RTP分布（Spin >= 100)')
    sht.cell(33, 14, '(0,0.3]')
    sht.cell(34, 14, '(0.3,0.4]')
    sht.cell(35, 14, '(0.4,0.5]')
    sht.cell(36, 14, '(0.5,0.6]')
    sht.cell(37, 14, '(0.6,0.7]')
    sht.cell(38, 14, '(0.7,0.8]')
    sht.cell(39, 14, '(0.8,0.9]')
    sht.cell(40, 14, '(0.9,1]')
    sht.cell(41, 14, '(1,1.1]')
    sht.cell(42, 14, '(1.1,1.2]')
    sht.cell(43, 14, '(1.2,1.5]')
    sht.cell(44, 14, '(1.5,2]')
    sht.cell(45, 14, '(2,max]')
    sht.cell(46, 14, '')
    sht.cell(47, 14, '人均RTP')
    sht.cell(48, 14, '')
    sht.cell(49, 14, '')
    sht.cell(50, 14, '累计Spin次数')
    sht.cell(51, 14, 'Spin人数')

    sht.cell(1, 15, '玩家分布')
    sht.cell(2, 15, '')
    sht.cell(3, 15, '=COUNTIF(B:B,0)')
    sht.cell(4, 15, '=COUNTIFS(B:B,">0",B:B,"<=10")')
    sht.cell(5, 15, '=COUNTIFS(B:B,">10",B:B,"<=50")')
    sht.cell(6, 15, '=COUNTIFS(B:B,">50",B:B,"<=100")')
    sht.cell(7, 15, '=COUNTIFS(B:B,">100",B:B,"<=500")')
    sht.cell(8, 15, '=COUNTIF(B:B,">500")')
    sht.cell(9, 15, '')
    sht.cell(10, 15, '=AVERAGE(B:B)')
    sht.cell(11, 15, '=AVERAGEIF(B:B,">=100",B:B)')
    sht.cell(12, 15, '=P7+P8')
    sht.cell(13, 15, '')
    sht.cell(14, 15, '玩家分布')
    sht.cell(15, 15, '=COUNTIFS(F:F,">=0",F:F,"<=0.3")')
    sht.cell(16, 15, '=COUNTIFS(F:F,">0.3",F:F,"<=0.4")')
    sht.cell(17, 15, '=COUNTIFS(F:F,">0.4",F:F,"<=0.5")')
    sht.cell(18, 15, '=COUNTIFS(F:F,">0.5",F:F,"<=0.6")')
    sht.cell(19, 15, '=COUNTIFS(F:F,">0.6",F:F,"<=0.7")')
    sht.cell(20, 15, '=COUNTIFS(F:F,">0.7",F:F,"<=0.8")')
    sht.cell(21, 15, '=COUNTIFS(F:F,">0.8",F:F,"<=0.9")')
    sht.cell(22, 15, '=COUNTIFS(F:F,">0.9",F:F,"<=1")')
    sht.cell(23, 15, '=COUNTIFS(F:F,">1",F:F,"<=1.1")')
    sht.cell(24, 15, '=COUNTIFS(F:F,">1.1",F:F,"<=1.2")')
    sht.cell(25, 15, '=COUNTIFS(F:F,">1.2",F:F,"<=1.5")')
    sht.cell(26, 15, '=COUNTIFS(F:F,">1.5",F:F,"<=2")')
    sht.cell(27, 15, '=COUNTIF(F:F,">2")')
    sht.cell(28, 15, '')
    sht.cell(29, 15, '=AVERAGE(F:F)')
    sht.cell(30, 15, '=SUM(E:E)/SUM(B:B)')
    sht.cell(31, 15, '')
    sht.cell(32, 15, '玩家分布')
    sht.cell(33, 15, '=COUNTIFS(B:B,">=100",F:F,"<=0.3")')
    sht.cell(34, 15, '=COUNTIFS(B:B,">=100",F:F,">0.3",F:F,"<=0.4")')
    sht.cell(35, 15, '=COUNTIFS(B:B,">=100",F:F,">0.4",F:F,"<=0.5")')
    sht.cell(36, 15, '=COUNTIFS(B:B,">=100",F:F,">0.5",F:F,"<=0.6")')
    sht.cell(37, 15, '=COUNTIFS(B:B,">=100",F:F,">0.6",F:F,"<=0.7")')
    sht.cell(38, 15, '=COUNTIFS(B:B,">=100",F:F,">0.7",F:F,"<=0.8")')
    sht.cell(39, 15, '=COUNTIFS(B:B,">=100",F:F,">0.8",F:F,"<=0.9")')
    sht.cell(40, 15, '=COUNTIFS(B:B,">=100",F:F,">0.9",F:F,"<=1.0")')
    sht.cell(41, 15, '=COUNTIFS(B:B,">=100",F:F,">1.0",F:F,"<=1.1")')
    sht.cell(42, 15, '=COUNTIFS(B:B,">=100",F:F,">1.1",F:F,"<=1.2")')
    sht.cell(43, 15, '=COUNTIFS(B:B,">=100",F:F,">1.2",F:F,"<=1.5")')
    sht.cell(44, 15, '=COUNTIFS(B:B,">=100",F:F,">1.5",F:F,"<=2.0")')
    sht.cell(45, 15, '=COUNTIFS(B:B,">=100",F:F,">2")')
    sht.cell(46, 15, '')
    sht.cell(47, 15, '=SUMIF(B:B,">=100",F:F)/COUNTIF(B:B,">=100")')
    sht.cell(48, 15, '')
    sht.cell(49, 15, '')
    sht.cell(50, 15, '=SUM(B:B)')
    sht.cell(51, 15, '=COUNT(B:B)')

    sht.cell(1, 16, '占比')
    sht.cell(2, 16, '')
    sht.cell(3, 16, '=O3/SUM($O$3:$O$8)')
    sht.cell(4, 16, '=O4/SUM($O$3:$O$8)')
    sht.cell(5, 16, '=O5/SUM($O$3:$O$8)')
    sht.cell(6, 16, '=O6/SUM($O$3:$O$8)')
    sht.cell(7, 16, '=O7/SUM($O$3:$O$8)')
    sht.cell(8, 16, '=O8/SUM($O$3:$O$8)')
    sht.cell(9, 16, '')
    sht.cell(10, 16, '')
    sht.cell(11, 16, '')
    sht.cell(12, 16, '')
    sht.cell(13, 16, '')
    sht.cell(14, 16, '占比')
    sht.cell(15, 16, '=O15/SUM($O$15:$O$27)')
    sht.cell(16, 16, '=O16/SUM($O$15:$O$27)')
    sht.cell(17, 16, '=O17/SUM($O$15:$O$27)')
    sht.cell(18, 16, '=O18/SUM($O$15:$O$27)')
    sht.cell(19, 16, '=O19/SUM($O$15:$O$27)')
    sht.cell(20, 16, '=O20/SUM($O$15:$O$27)')
    sht.cell(21, 16, '=O21/SUM($O$15:$O$27)')
    sht.cell(22, 16, '=O22/SUM($O$15:$O$27)')
    sht.cell(23, 16, '=O23/SUM($O$15:$O$27)')
    sht.cell(24, 16, '=O24/SUM($O$15:$O$27)')
    sht.cell(25, 16, '=O25/SUM($O$15:$O$27)')
    sht.cell(26, 16, '=O26/SUM($O$15:$O$27)')
    sht.cell(27, 16, '=O27/SUM($O$15:$O$27)')
    sht.cell(28, 16, '')
    sht.cell(29, 16, '')
    sht.cell(30, 16, '')
    sht.cell(31, 16, '')
    sht.cell(32, 16, '占比')
    sht.cell(33, 16, '=O33/SUM($O$33:$O$45)')
    sht.cell(34, 16, '=O34/SUM($O$33:$O$45)')
    sht.cell(35, 16, '=O35/SUM($O$33:$O$45)')
    sht.cell(36, 16, '=O36/SUM($O$33:$O$45)')
    sht.cell(37, 16, '=O37/SUM($O$33:$O$45)')
    sht.cell(38, 16, '=O38/SUM($O$33:$O$45)')
    sht.cell(39, 16, '=O39/SUM($O$33:$O$45)')
    sht.cell(40, 16, '=O40/SUM($O$33:$O$45)')
    sht.cell(41, 16, '=O41/SUM($O$33:$O$45)')
    sht.cell(42, 16, '=O42/SUM($O$33:$O$45)')
    sht.cell(43, 16, '=O43/SUM($O$33:$O$45)')
    sht.cell(44, 16, '=O44/SUM($O$33:$O$45)')
    sht.cell(45, 16, '=O45/SUM($O$33:$O$45)')


