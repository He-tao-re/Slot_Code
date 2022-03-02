import os
import openpyxl
import Data_Analyze.RTP_Static as RTP_Static
import Data_Analyze.User_Spin_Count as User_Spin

import Data_Analyze.RTP_Static.G10001 as RTP_Static_G10001
import Data_Analyze.RTP_Static.G10002 as RTP_Static_G10002
import Data_Analyze.RTP_Static.G10003 as RTP_Static_G10003
import Data_Analyze.RTP_Static.G10004 as RTP_Static_G10004
import Data_Analyze.RTP_Static.G10005 as RTP_Static_G10005
import Data_Analyze.RTP_Static.G10006 as RTP_Static_G10006
import Data_Analyze.RTP_Static.G10007 as RTP_Static_G10007
import Data_Analyze.RTP_Static.G10008 as RTP_Static_G10008
import Data_Analyze.RTP_Static.G10009 as RTP_Static_G10009
import Data_Analyze.RTP_Static.G10010 as RTP_Static_G10010
import Data_Analyze.RTP_Static.G10011 as RTP_Static_G10011
import Data_Analyze.RTP_Static.G10012 as RTP_Static_G10012
import Data_Analyze.RTP_Static.G10013 as RTP_Static_G10013
import Data_Analyze.RTP_Static.G10014 as RTP_Static_G10014
import Data_Analyze.RTP_Static.G10015 as RTP_Static_G10015
import Data_Analyze.RTP_Static.G10016 as RTP_Static_G10016
import Data_Analyze.RTP_Static.G10017 as RTP_Static_G10017
import Data_Analyze.RTP_Static.G10018 as RTP_Static_G10018
import Data_Analyze.RTP_Static.G10019 as RTP_Static_G10019
import Data_Analyze.RTP_Static.G10020 as RTP_Static_G10020
import Data_Analyze.RTP_Static.G10021 as RTP_Static_G10021
import Data_Analyze.RTP_Static.G10022 as RTP_Static_G10022
import Data_Analyze.RTP_Static.G10023 as RTP_Static_G10023
import Data_Analyze.RTP_Static.G10024 as RTP_Static_G10024
import Data_Analyze.RTP_Static.G10025 as RTP_Static_G10025
import Data_Analyze.RTP_Static.G10026 as RTP_Static_G10026
import Data_Analyze.RTP_Static.G10028 as RTP_Static_G10028

import Data_Analyze.RTP_Static.G10029 as RTP_Static_G10029


import Data_Analyze.RTP_Static.G10036 as RTP_Static_G10036
import xlwings as xw
import openpyxl
summary_data = '/Users/ht/Documents/Online_Data/Analyze_Data/Summary_Data.xlsx'

def test_rtp_data(Sub_Data,Sum_Data):
    #删除旧excel，新建excel
    try:
        os.remove(summary_data)
        wb = openpyxl.Workbook()
        wb.save(summary_data)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        wb.save(summary_data)
    print(10004)

    RTP_Static_G10004.static_data(summary_data,Sub_Data,Sum_Data)


def rtp_data(Sub_Data,Sum_Data):

    #删除旧excel，新建excel
    try:
        os.remove(summary_data)
        wb = openpyxl.Workbook()
        wb.save(summary_data)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        wb.save(summary_data)

    app = xw.App(visible=False, add_book=False)
    app.display_alerts = False
    app.screen_updating = False
    wb = app.books.add()
    wb.save(path=summary_data)
    wb.close()
    app.quit()

    # # 机台RTP数据统计
    print(10001)
    RTP_Static_G10001.static_data(summary_data,Sub_Data,Sum_Data)
    print(10002)
    RTP_Static_G10002.static_data(summary_data,Sub_Data,Sum_Data)
    print(10003)
    RTP_Static_G10003.static_data(summary_data,Sub_Data,Sum_Data)
    # RTP_Static_G10004.static_data(summary_data,Sub_Data,Sum_Data)
    print(10005)
    RTP_Static_G10005.static_data(summary_data,Sub_Data,Sum_Data)
    # # RTP_Static_G10006.static_data(summary_data,Sub_Data,Sum_Data)
    # # RTP_Static_G10007.static_data(summary_data,Sub_Data,Sum_Data)
    print(10008)
    RTP_Static_G10008.static_data(summary_data,Sub_Data,Sum_Data)
    print(10009)
    RTP_Static_G10009.static_data(summary_data,Sub_Data,Sum_Data)
    print(10010)
    RTP_Static_G10010.static_data(summary_data,Sub_Data,Sum_Data)
    print(10011)
    RTP_Static_G10011.static_data(summary_data,Sub_Data,Sum_Data)
    print(10012)
    RTP_Static_G10012.static_data(summary_data,Sub_Data,Sum_Data)
    print(10013)
    RTP_Static_G10013.static_data(summary_data,Sub_Data,Sum_Data)
    print(10014)
    RTP_Static_G10014.static_data(summary_data,Sub_Data,Sum_Data)
    print(10015)
    RTP_Static_G10015.static_data(summary_data,Sub_Data,Sum_Data)
    print(10016)
    RTP_Static_G10016.static_data(summary_data,Sub_Data,Sum_Data)
    print(10017)
    RTP_Static_G10017.static_data(summary_data,Sub_Data,Sum_Data)
    print(10018)
    RTP_Static_G10018.static_data(summary_data,Sub_Data,Sum_Data)
    print(10019)
    RTP_Static_G10019.static_data(summary_data,Sub_Data,Sum_Data)
    print(10020)
    RTP_Static_G10020.static_data(summary_data,Sub_Data,Sum_Data)
    print(10021)
    RTP_Static_G10021.static_data(summary_data,Sub_Data,Sum_Data)
    print(10022)
    RTP_Static_G10022.static_data(summary_data,Sub_Data,Sum_Data)
    print(10024)
    RTP_Static_G10024.static_data(summary_data,Sub_Data,Sum_Data)
    print(10025)
    RTP_Static_G10025.static_data(summary_data,Sub_Data,Sum_Data)
    print(10026)
    RTP_Static_G10026.static_data(summary_data,Sub_Data,Sum_Data)
    print(10028)
    RTP_Static_G10028.static_data(summary_data,Sub_Data,Sum_Data)
    print(10029)
    RTP_Static_G10029.static_data(summary_data,Sub_Data,Sum_Data)
    print(10036)
    RTP_Static_G10036.static_data(summary_data,Sub_Data,Sum_Data)



def final_data(Key):
    if Key is True:
        wb = openpyxl.load_workbook(summary_data)
        sheet_list = wb.sheetnames
        sheet_list.remove("Sheet1")

        write_sheet = wb['Sheet1']
        write_sheet.cell(1,1,'Game_ID')
        write_sheet.cell(1,2,'人均Spin')
        write_sheet.cell(1,3,'人均spin（Spin >= 100）')
        write_sheet.cell(1,4,'深度体验率')
        write_sheet.cell(1,5,'人均RTP')
        write_sheet.cell(1,6,'关卡RTP')
        write_sheet.cell(1,7,'人均RTP（Spin >= 100）')
        write_sheet.cell(1,8,'累计Spin次数')
        write_sheet.cell(1,9,'Spin人数')

        row_num = 2
        for game_id in sheet_list:
            write_sheet.cell(row_num,1,int(game_id))
            write_sheet.cell(row_num,2,f'={game_id}!O10')
            write_sheet.cell(row_num,3,f'={game_id}!O11')
            write_sheet.cell(row_num,4,f'={game_id}!O12')
            write_sheet.cell(row_num,5,f'={game_id}!O29')
            write_sheet.cell(row_num,6,f'={game_id}!O30')
            write_sheet.cell(row_num,7,f'={game_id}!O47')
            write_sheet.cell(row_num,8,f'={game_id}!O50')
            write_sheet.cell(row_num,9,f'={game_id}!O51')
            row_num += 1

        # 玩家Spin数据统计
        write_sheet_2 = wb.create_sheet('User_Spin')

        uid_list = sorted(User_Spin.User_Spin_Count.keys())

        game_list = [i for i in range(10001,10050)]

        write_sheet_2.cell(1, 1, "uid")
        write_sheet_2.cell(1, 2, "Sum")

        for game_id_idx in range(len(game_list)):
            write_sheet_2.cell(1, game_id_idx+3, game_list[game_id_idx])

        for uid_idx in range(len(uid_list)):
            uid = uid_list[uid_idx]
            write_sheet_2.cell(uid_idx+2,1,uid)
            write_sheet_2.cell(uid_idx+2,2,f"=sum(C{uid_idx+2}:AZ{uid_idx+2})")

            for k,v in User_Spin.User_Spin_Count[uid].items():
                col_idx = game_list.index(k) + 3
                write_sheet_2.cell(uid_idx+2,col_idx,v)

        wb.save(summary_data)
