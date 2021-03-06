import json
import os
import csv
import time
import shutil
import openpyxl

Game_ID = 10013

"""File"""

Person_File_Site = "/Users/ht/Documents/Online_Data/Games_Users_Data_2/10013/7062.txt"
Analyze_Data_Site = "/Users/ht/Documents/7062.xlsx"


def write_down_excel():
    wb = openpyxl.Workbook()
    wb.save(Analyze_Data_Site)
    wb = openpyxl.load_workbook(Analyze_Data_Site)

    sht = wb.create_sheet("data")

    title = ["时间", "game_type", "uid", "user_level", "持金", "spinTimes", "rtp_type", "送奖", "bet", "win", "倍数", "free进度"]

    for i in range(len(title)):
        i = i + 1
        sht.cell(1, i, title[i-1])
    num = 1

    O_file = open(Person_File_Site,'r', encoding='UTF-8',newline='')
    for spin_data in O_file:

        # print(spin_data)

        spin_data = json.loads(spin_data)

        game_type = spin_data["data"]['gameType']
        uid = spin_data["data"]['uid']
        user_level = spin_data["data"]["gradeData"]["level"]
        balance_gold = spin_data["data"]["gold"]

        spin_times = spin_data["data"]["attachInfo"]["testModel"]["spinTimes"]

        if "currentRtp" in spin_data["data"]["attachInfo"]["testModel"].keys():
            rtp_type = spin_data["data"]["attachInfo"]["testModel"]["currentRtp"]
        else:
            rtp_type = ""

        if "rewardId" in spin_data["data"]["attachInfo"]["testModel"].keys():
            reward_type = spin_data["data"]["attachInfo"]["testModel"]["rewardId"]
        else:
            reward_type = ""


        bet = spin_data["data"]['betGold']
        win = spin_data["data"]['winGold']

        super_free = spin_data["data"]['superFreeGameTimes']

        """时间格式"""
        timeStamp = float(spin_data['data']['serverTime']/1000)
        timeArray = time.localtime(timeStamp)
        otherStyleTime = time.strftime("%Y-%m-%d %H:%M:%S", timeArray)



        num += 1
        spin_mul = win / bet

        data_1 = [otherStyleTime, game_type, uid, user_level, balance_gold, spin_times, rtp_type, reward_type, bet, win, spin_mul, super_free]

        for i in range(len(data_1)):
            i = i + 1
            sht.cell(num, i, data_1[i - 1])

    wb.save(Analyze_Data_Site)




if __name__ == '__main__':
    write_down_excel()